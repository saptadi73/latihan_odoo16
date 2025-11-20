from odoo import api, fields, models, _
from odoo.exceptions import UserError, ValidationError

import base64
from io import StringIO
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import logging
import pathlib
import os

_log = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Border, Font, Alignment
except ImportError:
    _log.debug('Can not `import openpyxl`.')

class KSWarehouseReportAgeingNoMovements(models.Model):
    _name = "ks.warehouse.report.ageing.no.movement"
    _description = "Stock Ageing No Movements / Stock Report"

    ks_report = {'product_code': 0, 'product_type': 1, 'product_categ_id': 2, 'product_name': 3, 'location_id': 4,
                 'company_id': 5, 'product_sales_price': 6, 'product_qty_available': 7, 'product_id': 8}
    kr_in_dates = {'product_id': 0, 'location_id': 1, 'company_id': 2, 'opening_stock': 3, 'closing_stock': 4,
                   'qty_date': 5}

    ks_name = fields.Char(default='Stock Aging No Movement Report')
    ks_date_from = fields.Date('Start Date', required=True)
    ks_date_to = fields.Date('End Date', required=True)
    ks_company_id = fields.Many2one('res.company', 'Company', required=True,
                                 default=lambda self: self.env.user.company_id)
    ks_duration = fields.Integer('Duration Range', required=True, default=30)
    
    def ks_action_generate_report(self):
        return True

    def ks_apply_style(self, ks_cell, kc='', vc='', sz=False, wp=False):
        ks_cell.alignment = Alignment(horizontal="center" if kc else '', vertical="center" if vc else '',
                                      wrap_text=wp)
        if sz: ks_cell.font = Font(b=True, size=sz)

    def ks_set_default_5_columns_left(self, report_name, sheet):
        sheet.title = str(report_name)

        sheet['A1'] = str(report_name)
        self.ks_apply_style(sheet['A1'], True, True, 20, True)
        sheet.merge_cells(start_row=1, end_row=2, start_column=1, end_column=20)

        sheet['A3'] = "COMPANY : " + self.ks_company_id.name
        self.ks_apply_style(sheet['A3'], True, True, 14, True)
        sheet.merge_cells(start_row=3, end_row=3, start_column=1, end_column=20)

        sheet['A4'] = 'FROM : ' + str(self.ks_date_from) + ' | TO : ' + str(self.ks_date_to)
        self.ks_apply_style(sheet['A4'], True, True, 10, True)
        sheet.merge_cells(start_row=4, end_row=4, start_column=1, end_column=20)

        sheet['A6'] = "REPORT"
        self.ks_apply_style(sheet['A6'], True, True, 14, True)
        sheet.merge_cells(start_row=6, end_row=7, start_column=1, end_column=20)

        sheet['A8'] = "S.NO"
        self.ks_apply_style(sheet['A8'], True, True, False, True)
        sheet.merge_cells(start_row=8, end_row=10, start_column=1, end_column=1)

        sheet['B8'] = "Reference/Code"
        self.ks_apply_style(sheet['B8'], True, True, False, True)
        sheet.merge_cells(start_row=8, end_row=10, start_column=2, end_column=2)

        sheet['C8'] = "Type"
        self.ks_apply_style(sheet['C8'], True, True, False, True)
        sheet.merge_cells(start_row=8, end_row=10, start_column=3, end_column=3)

        sheet['D8'] = "Category"
        self.ks_apply_style(sheet['D8'], True, True, False, True)
        sheet.merge_cells(start_row=8, end_row=10, start_column=4, end_column=4)

        sheet['E8'] = "Product"
        self.ks_apply_style(sheet['E8'], True, True, False, True)
        sheet.merge_cells(start_row=8, end_row=10, start_column=5, end_column=5)

        sheet['F8'] = "Location"
        self.ks_apply_style(sheet['F8'], True, True, False, True)
        sheet.merge_cells(start_row=8, end_row=10, start_column=6, end_column=6)

        sheet.freeze_panes = 'C11'

    def ks_generate_xlsx_report(self):
        report_name = self.ks_name
        workbook = openpyxl.Workbook()

        sheet = workbook.active

        self.ks_set_default_5_columns_left(report_name, sheet)

        # get ranges
        period_length = self.ks_duration
        if period_length <= 0:
            raise UserError(_('You must set a period length greater than 0.'))
        if not self.ks_date_from:
            raise UserError(_('You must set a start date.'))
        ks_date_from = fields.Datetime.to_datetime(self.ks_date_from)
        ks_date_to = fields.Datetime.to_datetime(self.ks_date_to)

        ks_inc = 0
        ks_limit_days = (self.ks_date_to - self.ks_date_from).days
        for ks in range(0, ks_limit_days, self.ks_duration):
            if ks + self.ks_duration > ks_limit_days:
                rang = str(ks + 1) + "-" + str((self.ks_date_to - self.ks_date_from).days)
            else:
                rang = (str(ks) + "-" + str(ks + self.ks_duration))
            if self.ks_date_from + timedelta(days=ks + self.ks_duration) > self.ks_date_to:
                date_rang = str((self.ks_date_from + timedelta(days=ks)).strftime('%d-%m-%Y')) + ':' +\
                            str((self.ks_date_to - timedelta(days=1)).strftime('%d-%m-%Y'))
            else:
                date_rang = str((self.ks_date_from + timedelta(days=ks)).strftime('%d-%m-%Y')) + ':' + \
                            str((self.ks_date_from + timedelta(days=ks + self.ks_duration - 1)).strftime('%d-%m-%Y'))

            # name = str(self.ks_date_from + timedelta(days=i))
            # if 6 + ks_inc + 2 > 255: # It will enter the loop if column greater than 255 to split the sheet into new
            #     ks_inc = 0
            #     if 'ks_sheet' not in locals(): ks_sheet = {}
            #     if 'n_sheet' in locals(): n_sheet += 1
            #     if 'n_sheet' not in locals():
            #         n_sheet = 1
            #         ks_sheet[n_sheet] = sheet
            #     sheet = workbook.create_sheet(str(n_sheet) + '_' + str(report_name))
            #     self.ks_set_default_5_columns_left(report_name, sheet)
            #     ks_sheet[n_sheet + 1] = sheet

            sheet.cell(8, 7 + ks_inc, rang)
            sheet.cell(9, 7 + ks_inc, date_rang)
            sheet.cell(10, 7 + ks_inc, "Stock Qty")
            sheet.cell(10, 8 + ks_inc, "Stock Value")
            sheet.merge_cells(start_row=8, end_row=8, start_column=7 + ks_inc, end_column=8 + ks_inc)
            sheet.merge_cells(start_row=9, end_row=9, start_column=7 + ks_inc, end_column=8 + ks_inc)
            ks_inc += 2
        # else:
        #     if 'ks_sheet' in locals():
        #         sheet =ks_sheet[1]

        # get qty available
        self.env.cr.execute("""
            SELECT ks_product_id, ks_product_type,ks_product_categ_id,ks_product_name,ks_location_id,ks_company_id,
                   ks_product_sales_price, ks_product_qty_available, ks_product_code
            FROM ks_warehouse_report 
            WHERE ks_company_id = %s and ks_product_qty_available != 0 and ks_usage = 'internal'
            order by ks_location_id
        """ % self.ks_company_id.id)

        datas = self.env.cr.fetchall()
        if not datas:
            raise ValidationError(_("Opps! There are no data."))

        # dates_in = self.ks_data_in_date(periods)

        # datas = self.ks_merge_data(datas, dates_in)

        # datas = self.ks_apply_filter(datas)

        if datas:
            j = 1;
            i = 1;
            row = 11;
            col = 0
            for data in datas:
                self.ks_set_default_5_rows_left(sheet, row, j, data)
                ks_col = 0
                start = datetime.strptime(str(self.ks_date_from), "%Y-%m-%d")
                for i in range(0, (self.ks_date_to - self.ks_date_from).days, self.ks_duration):
                    # if 6 + ks_col + 2 > 255:  # It will enter the loop if column gretaer than 255 to split the sheet into new
                    #     ks_col = 0
                    #     if 'nd_sheet' in locals(): nd_sheet += 1
                    #     if 'nd_sheet' not in locals(): nd_sheet = 2
                    #     sheet = ks_sheet[nd_sheet]
                    #     self.ks_set_default_5_rows_left(sheet, row, j, data)

                    stop = start + relativedelta(days=period_length)
                    stop = ks_date_to if stop > ks_date_to else stop

                    self.env.cr.execute("""
                                          select sq.product_id, sq.company_id,
                                          sum(case when sq.create_date >= '%s' then  sq.quantity else 0 end) as qty_date
                                          from stock_valuation_layer as sq
                                              LEFT JOIN stock_move as sm on sm.id = sq.stock_move_id
                                              LEFT JOIN stock_location as sl ON sl.id = sm.location_id
                                          where sq.product_id = '%s'
                                              and sq.company_id = '%s'
                                              and sq.create_date <= '%s'
                                              and sm.location_id = '%s'
                                          group by sq.product_id, sq.company_id
                                          """ % (
                        start, data[0], self.ks_company_id.id, fields.Datetime.to_datetime(stop), data[4]))
                    pro_data = self.env.cr.fetchall()
                    sheet.cell(row, 7+ks_col, pro_data[0][2] if pro_data else 0)
                    ks_cost = self.env['product.product'].browse(data[0]).product_tmpl_id.standard_price
                    sheet.cell(row, 8+ks_col, (pro_data[0][2] if pro_data else 0)*ks_cost)
                    start = stop
                    ks_col += 2
                # else:
                #     if 'ks_sheet' in locals(): sheet = ks_sheet[1]
                #     if 'nd_sheet' in locals(): nd_sheet = 1

                row += 1
                j += 1
        output = StringIO()
        filename = (str(pathlib.Path(__file__).parent) + '/' + str(report_name) + '.xlsx')
        _log.info("Filepath: %s" % (filename))
        workbook.save(filename)
        fp = open(filename, "rb")
        file_data = fp.read()
        out = base64.encodebytes(file_data)

        # Files actions
        attach_vals = {
            'report_name': str(report_name) + '.xlsx',
            'datas': out,
        }

        act_id = self.env['ks.warehouse.report.ageing.no.movement.out'].create(attach_vals)
        fp.close()
        os.remove(filename)
        _log.info("File closed and removed.")
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'ks.warehouse.report.ageing.no.movement.out',
            'res_id': act_id.id,
            'view_type': 'form',
            'view_mode': 'form',
            'context': self.env.context,
            'target': 'new',
        }

    
    def ks_set_default_5_rows_left(self, sheet, row, j, data):
        sheet.cell(row, 1, j)
        sheet.cell(row, 2, data[0])
        if data[1] == 'product':
            sheet.cell(row, 3, 'Stockable')
        elif data[1] == 'consu':
            sheet.cell(row, 3, 'Consumable')
        if data[2]:
            catge_id = self.env['product.category'].browse(int(data[2]))
        sheet.cell(row, 4, catge_id.name)
        sheet.cell(row, 5, data[3])
        if data[4]:
            location_id = self.env['stock.location'].browse(int(data[4]))
        sheet.cell(row, 6, location_id.display_name)

    class KSWarehouseReportAgeingNoMovementOUT(models.Model):
        _name = "ks.warehouse.report.ageing.no.movement.out"
        _description = "Stock Ageing No Movement report Out"

        datas = fields.Binary('File', readonly=True)
        report_name = fields.Char('Report Name', readonly=True)
