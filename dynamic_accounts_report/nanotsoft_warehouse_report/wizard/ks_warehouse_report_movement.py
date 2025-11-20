from odoo import api, fields, models, _
from odoo.exceptions import ValidationError

import base64
from io import StringIO
import logging
import pathlib
import os

_log = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Border, Font, Alignment
except ImportError:
    _log.debug('Can not `import openpyxl`.')


class KSWarehouseReportmovements(models.Model):
    _name = "ks.warehouse.report.movement"
    _description = "Stock movements / Stock Report"

    ks_report = {'product_code': 0, 'product_type': 1, 'product_categ_id': 2, 'product_name': 3, 'location_id': 4,
                 'company_id': 5, 'product_sales_price': 6, 'product_qty_available': 7, 'product_id': 8}
    ks_operate = {'product_id': 0, 'location_id': 1, 'company_id': 2, 'scheduled_date': 3, 'p_type_id': 4,
                  'location_id': 5, 'location_dest_id': 6, 'p_uom_qty': 7, 'user_id': 8, 'product_name': 9,
                  'product_type': 10, 'product_categ_id': 11}

    ks_name = fields.Char(default='Stock Movement Report')
    ks_date_from = fields.Date('Start Date', required=True)
    ks_date_to = fields.Date('End Date', required=True)
    ks_company_id = fields.Many2one('res.company', 'Company', required=True,
                                    default=lambda self: self.env.user.company_id)
    ks_operation_type = fields.Many2many('stock.picking.type', string='Operation Types')
    ks_responsible = fields.Many2many('res.users', string='Responsible User')
    ks_category = fields.Many2many('product.category', string='Product Category')

    def ks_action_generate_report(self):
        return True

    def ks_apply_style(self, ks_cell, kc='', vc='', sz=False, wp=False):
        ks_cell.alignment = Alignment(horizontal="center" if kc else '', vertical="center" if vc else '',
                                      wrap_text=wp)
        if sz: ks_cell.font = Font(b=True, size=sz)

    def ks_create_workbook_header(self, report_name, sheet):
        sheet.title = str(report_name)

        sheet['A1'] = str(report_name)
        self.ks_apply_style(sheet['A1'], True, True, 20, True)
        sheet.merge_cells(start_row=1, end_row=2, start_column=1, end_column=20)

        sheet['A3'] = "COMPANY : " + self.ks_company_id.name
        self.ks_apply_style(sheet['A3'], True, True, 14, True)
        sheet.merge_cells(start_row=3, end_row=3, start_column=1, end_column=20)

        sheet['A4'] = 'FROM : ' + str(self.ks_date_from) + ' | TO : ' + str(self.ks_date_to)
        self.ks_apply_style(sheet['A4'], True, True, 10, True)
        sheet.merge_cells(start_row=4, end_row=4, start_column=1, end_column=20)

        sheet['A6'] = "REPORT"
        self.ks_apply_style(sheet['A6'], True, True, 14, True)
        sheet.merge_cells(start_row=6, end_row=7, start_column=1, end_column=20)

        sheet['A8'] = "S.NO"
        sheet.merge_cells(start_row=8, end_row=9, start_column=1, end_column=1)

        sheet['B8'] = "Reference/Code"
        self.ks_apply_style(sheet['B8'], True, True, False, True)
        sheet.merge_cells(start_row=8, end_row=9, start_column=2, end_column=2)

        sheet['C8'] = "Type"
        sheet.merge_cells(start_row=8, end_row=9, start_column=3, end_column=3)

        sheet['D8'] = "Category"
        sheet.merge_cells(start_row=8, end_row=9, start_column=4, end_column=4)

        sheet['E8'] = "Product"
        sheet.merge_cells(start_row=8, end_row=9, start_column=5, end_column=5)

        # sheet['F8'] = "Location"
        # sheet.merge_cells(start_row=8, end_row=9, start_column=6, end_column=6)

        sheet['F8'] = "Company"
        sheet.merge_cells(start_row=8, end_row=9, start_column=6, end_column=6)

        sheet['G8'] = "Inventory Date"
        self.ks_apply_style(sheet['G8'], True, True, False, True)
        sheet.merge_cells(start_row=8, end_row=9, start_column=7, end_column=7)

        sheet['H8'] = "Operation Types"
        self.ks_apply_style(sheet['H8'], True, True, False, True)
        sheet.merge_cells(start_row=8, end_row=9, start_column=8, end_column=8)

        sheet['I8'] = "Locations"
        self.ks_apply_style(sheet['I8'], True, True, False, True)
        sheet['I9'] = "Source"
        sheet['J9'] = "Destination"
        sheet.merge_cells(start_row=8, end_row=8, start_column=9, end_column=10)

        sheet['K8'] = "Qty"
        sheet.merge_cells(start_row=8, end_row=9, start_column=11, end_column=11)

        sheet['L8'] = "Responsibles"
        self.ks_apply_style(sheet['L8'], True, True, False, True)
        sheet.merge_cells(start_row=8, end_row=9, start_column=12, end_column=12)

        sheet.freeze_panes = 'C10'

    def ks_generate_xlsx_report(self):
        report_name = self.ks_name
        workbook = openpyxl.Workbook()

        sheet = workbook.active

        self.ks_create_workbook_header(report_name, sheet)

        # # get qty available
        # self.env.cr.execute("""
        # SELECT ks_product_code,ks_product_type,ks_product_categ_id,ks_product_name,ks_location_id,ks_company_id,
        #        ks_product_sales_price, ks_product_qty_available, ks_product_id
        # FROM ks_warehouse_report
        # WHERE ks_company_id = %s and ks_product_qty_available != 0 and ks_usage = 'internal'
        # order by ks_location_id
        # """ % self.ks_company_id.id)
        #
        # datas = self.env.cr.fetchall()
        # if not datas:
        #     raise ValidationError(_("Opps! There are no data."))

        ks_operations = self.ks_operations()

        datas = self.ks_merge_data(ks_operations)

        if datas:
            i = 1;
            row = 10;
            for data in datas:
                sheet.cell(row, 1, i)
                sheet.cell(row, 2, data[0])
                if data[1] == 'product':
                    sheet.cell(row, 3, 'Stockable')
                elif data[1] == 'consu':
                    sheet.cell(row, 3, 'Consumable')
                if data[2]:
                    catge_id = self.env['product.category'].browse(int(data[2]))
                sheet.cell(row, 4, catge_id.name)
                sheet.cell(row, 5, data[3])
                # if data[4]:
                #     location_id = self.env['stock.location'].browse(int(data[4]))
                # sheet.cell(row, 6, location_id.display_name)
                if data[5]:
                    comp_id = self.env['res.company'].browse(int(data[5]))
                sheet.cell(row, 6, comp_id.name)
                sheet.cell(row, 7, data[6])
                if data[7]:
                    ks_pick = self.env['stock.picking.type'].browse(int(data[7])) if int(data[7]) > 0 else {'code': ''}
                sheet.cell(row, 8, ks_pick.code)
                if data[8]:
                    location_id = self.env['stock.location'].browse(int(data[8])) if int(data[8]) > 0 else {'name': ''}
                sheet.cell(row, 9, location_id.display_name)
                if data[9]:
                    location_dest_id = self.env['stock.location'].browse(int(data[9])) if int(data[9]) > 0 else {'display_name': ''}
                sheet.cell(row, 10, location_dest_id.display_name)
                sheet.cell(row, 11, data[10])
                if data[11]:
                    user = self.env['res.users'].browse(int(data[11])) if int(data[11]) > 0 else {'name': ''}
                sheet.cell(row, 12, user.name)

                row += 1
                i += 1
        output = StringIO()
        filename = (str(pathlib.Path(__file__).parent) + '/' + str(report_name) + '.xlsx')
        _log.info("Filepath: %s" % (filename))
        workbook.save(filename)
        fp = open(filename, "rb")
        file_data = fp.read()
        out = base64.encodebytes(file_data)

        # Files actions
        attach_vals = {
            'report_name': str(report_name) + '.xlsx',
            'datas': out,
        }

        act_id = self.env['ks.warehouse.report.movement.out'].create(attach_vals)
        fp.close()
        os.remove(filename)
        _log.info("File closed and removed.")
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'ks.warehouse.report.movement.out',
            'res_id': act_id.id,
            'view_type': 'form',
            'view_mode': 'form',
            'context': self.env.context,
            'target': 'new',
        }


    def ks_operations(self):
        ks_date_from = fields.Datetime.to_datetime(self.ks_date_from)
        _select = """
            sm.product_id, sm.location_id, sm.company_id, sp.scheduled_date, sm.picking_type_id, sm.location_id, 
            sm.location_dest_id, sum(sm.product_uom_qty), sm.create_uid, max(pt.name) as ks_product_name, 
            max(pt.type) as ks_product_type, max(pc.id) as ks_product_categ_id
        """
        _from = """
            stock_move as sm
            left join stock_picking as sp on sp.id = sm.picking_id
            LEFT JOIN product_product as pp ON pp.id = sm.product_id
            LEFT JOIN product_template as pt ON pt.id = pp.product_tmpl_id
            LEFT JOIN product_category as pc ON pc.id = pt.categ_id
        """
        _where = """
            sm.state not in ('done', 'cancel') and sm.company_id = '%s' and 
            sp.scheduled_date between '%s' and '%s'
        """ % (self.ks_company_id.id, ks_date_from, fields.Datetime.to_datetime(self.ks_date_to))
        _groupby = """
            sm.product_id, sm.location_id, sm.location_dest_id, sm.company_id, sm.picking_type_id, sp.scheduled_date,
            sm.create_uid
        """

        if self.ks_category:
            if len(self.ks_category.ids) > 1:
                _where += """ and pc.id in %s""" % (tuple(self.ks_category.ids),)
            else:
                _where += """ and pc.id = %s""" % (tuple(self.ks_category.ids))
        if self.ks_responsible:
            if len(self.ks_responsible.ids) > 1:
                _where += """ and sm.create_uid in %s""" % (tuple(self.ks_responsible.ids),)
            else:
                _where += """ and sm.create_uid = %s""" % (tuple(self.ks_responsible.ids))
        if self.ks_operation_type:
            if len(self.ks_operation_type.ids) > 1:
                _where += """ and sm.picking_type_id in %s""" % (tuple(self.ks_operation_type.ids),)
            else:
                _where += """ and sm.picking_type_id = %s""" % (tuple(self.ks_operation_type.ids))

        self.env.cr.execute("""SELECT %s FROM %s WHERE %s GROUP BY %s
            """ % (_select, _from, _where, _groupby)
                            )

        ks_operations = self.env.cr.fetchall()
        if not ks_operations:
            raise ValidationError(_("Opps! There are no data."))
        return ks_operations


    def ks_merge_data(self, ks_operations):
        ks_list = []
        kr = self.ks_report
        kid = self.ks_operate
        for date in ks_operations:
            # for data in datas:
            #     if data[kr['product_id']] == date[kid['product_id']] \
            #             and data[kr['company_id']] == date[kid['company_id']]:
                    ks_list.append(
                (date[kid['product_id']], date[kid['product_type']], date[kid['product_categ_id']],
                 date[kid['product_name']], date[kid['location_id']], date[kid['company_id']],
                         date[kid['scheduled_date']].date().strftime('%d-%m-%Y'), date[kid['p_type_id']],
                         date[kid['location_id']], date[kid['location_dest_id']], date[kid['p_uom_qty']],
                         date[kid['user_id']]
                         )
                    )
        if not ks_list:
            raise ValidationError(_("Opps! There are no data."))
        return ks_list

    # def ks_apply_filter(self, data):
    #         # ks_data = self.ks_exhausted_filter(data)
    #         ks_data = data
    #         if not ks_data:
    #             raise ValidationError(_("Opps! There are no data."))
    #         return ks_data

    class KSWarehouseReportmovementOUT(models.Model):
        _name = "ks.warehouse.report.movement.out"
        _description = "Stock movement report Out"

        datas = fields.Binary('File', readonly=True)
        report_name = fields.Char('Report Name', readonly=True)
