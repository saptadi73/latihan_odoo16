from odoo import api, fields, models, _
from odoo.exceptions import UserError, ValidationError

import base64
from io import StringIO
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import logging
import pathlib
import os

_log = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Border, Font, Alignment
except ImportError:
    _log.debug('Can not `import openpyxl`.')

class KSWarehouseReportAgeingWithMovement(models.Model):
    _name = "ks.warehouse.report.ageing.with.movement"
    _description = "Stock Ageing With Movements / Stock Report"

    ks_report = {'product_code': 0, 'product_type': 1, 'product_categ_id': 2, 'product_name': 3, 'location_id': 4,
                 'company_id': 5, 'product_sales_price': 6, 'product_qty_available': 7, 'product_id': 8}
    kr_in_dates = {'product_id': 0, 'location_id': 1, 'company_id': 2, 'opening_stock': 3, 'closing_stock': 4,
                   'qty_date': 5}

    ks_name = fields.Char(default='Stock Aging With Movement Report')
    ks_date_from = fields.Date('Start Date', required=True)
    ks_date_to = fields.Date('End Date', required=True)
    ks_company_id = fields.Many2one('res.company', 'Company', required=True,
                                 default=lambda self: self.env.user.company_id)
    ks_duration = fields.Integer('Duration Range', required=True, default=30)

    def ks_action_generate_report(self):
        return True

    def ks_apply_style(self, ks_cell, kc='', vc='', sz=False, wp=False):
        ks_cell.alignment = Alignment(horizontal="center" if kc else '', vertical="center" if vc else '',
                                      wrap_text=wp)
        if sz: ks_cell.font = Font(b=True, size=sz)

    def ks_set_default_5_columns_left(self, report_name, sheet):
        sheet.title = str(report_name)

        sheet['A1'] = str(report_name)
        self.ks_apply_style(sheet['A1'], True, True, 20, True)
        sheet.merge_cells(start_row=1, end_row=2, start_column=1, end_column=20)

        sheet['A3'] = "COMPANY : " + self.ks_company_id.name
        self.ks_apply_style(sheet['A3'], True, True, 14, True)
        sheet.merge_cells(start_row=3, end_row=3, start_column=1, end_column=20)

        sheet['A4'] = 'FROM : ' + str(self.ks_date_from) + ' | TO : ' + str(self.ks_date_to)
        self.ks_apply_style(sheet['A4'], True, True, 10, True)
        sheet.merge_cells(start_row=4, end_row=4, start_column=1, end_column=20)

        sheet['A6'] = "REPORT"
        self.ks_apply_style(sheet['A6'], True, True, 14, True)
        sheet.merge_cells(start_row=6, end_row=7, start_column=1, end_column=20)

        sheet['A8'] = "S.NO"
        self.ks_apply_style(sheet['A8'], True, True, False, True)
        sheet.merge_cells(start_row=8, end_row=10, start_column=1, end_column=1)

        sheet['B8'] = "Reference/Code"
        self.ks_apply_style(sheet['B8'], True, True, False, True)
        sheet.merge_cells(start_row=8, end_row=10, start_column=2, end_column=2)

        sheet['C8'] = "Type"
        self.ks_apply_style(sheet['C8'], True, True, False, True)
        sheet.merge_cells(start_row=8, end_row=10, start_column=3, end_column=3)

        sheet['D8'] = "Category"
        self.ks_apply_style(sheet['D8'], True, True, False, True)
        sheet.merge_cells(start_row=8, end_row=10, start_column=4, end_column=4)

        sheet['E8'] = "Product"
        self.ks_apply_style(sheet['E8'], True, True, False, True)
        sheet.merge_cells(start_row=8, end_row=10, start_column=5, end_column=5)

        sheet['F8'] = "Location"
        self.ks_apply_style(sheet['F8'], True, True, False, True)
        sheet.merge_cells(start_row=8, end_row=9, start_column=6, end_column=6)

        sheet.freeze_panes = 'C12'
    
    def ks_generate_xlsx_report(self):
        report_name = self.ks_name
        workbook = openpyxl.Workbook()

        sheet = workbook.active

        self.ks_set_default_5_columns_left(report_name, sheet)

        # get ranges
        period_length = self.ks_duration
        if period_length <= 0:
            raise UserError(_('You must set a period length greater than 0.'))
        if not self.ks_date_from:
            raise UserError(_('You must set a start date.'))
        ks_date_from = fields.Datetime.to_datetime(self.ks_date_from)
        ks_date_to = fields.Datetime.to_datetime(self.ks_date_to)

        ks_inc = 0
        ks_limit_days = (self.ks_date_to - self.ks_date_from).days
        for ks in range(0, ks_limit_days, self.ks_duration):
            if ks + self.ks_duration > ks_limit_days:
                rang = str(ks + 1) + "-" + str((self.ks_date_to - self.ks_date_from).days)
            else:
                rang = (str(ks) + "-" + str(ks + self.ks_duration)) # if ks == 0 else (str(ks+1) + "-" + str(ks + self.ks_duration))
            if self.ks_date_from + timedelta(days=ks + self.ks_duration) > self.ks_date_to:
                date_rang = str((self.ks_date_from + timedelta(days=ks)).strftime('%d-%m-%Y')) + ':' +\
                            str((self.ks_date_to - timedelta(days=1)).strftime('%d-%m-%Y'))
            else:
                date_rang = str((self.ks_date_from + timedelta(days=ks)).strftime('%d-%m-%Y')) + ':' + \
                            str((self.ks_date_from + timedelta(days=ks + self.ks_duration - 1)).strftime('%d-%m-%Y'))

            # if 14 + ks_inc > 255:  # It will enter the loop if column greater than 255 to split the sheet into new
            #     ks_inc = 0
            #     if 'ks_sheet' not in locals(): ks_sheet = {}
            #     if 'n_sheet' in locals(): n_sheet += 1
            #     if 'n_sheet' not in locals():
            #         n_sheet = 1
            #         ks_sheet[n_sheet] = sheet
            #     sheet = workbook.add_sheet(str(n_sheet) + '_' + str(report_name))
            #     self.ks_set_default_5_columns_left(sheet, report_name, header, header_small, column_style)
            #     ks_sheet[n_sheet + 1] = sheet

            sheet.cell(8, 7 + ks_inc, rang)
            self.ks_apply_style(sheet.cell(8, 7 + ks_inc), True, True, False, True)
            sheet.cell(9, 7 + ks_inc, date_rang)
            self.ks_apply_style(sheet.cell(9, 7 + ks_inc), True, True, False, True)
            sheet.merge_cells(start_row=8, end_row=8, start_column=7 + ks_inc, end_column=16 + ks_inc)
            sheet.merge_cells(start_row=9, end_row=9, start_column=7 + ks_inc, end_column=16 + ks_inc)
            sheet.cell(10, 7 + ks_inc, 'Sale')
            self.ks_apply_style(sheet.cell(10, 7 + ks_inc), True, True, False, True)
            sheet.cell(11, 7 + ks_inc, "Qty")
            sheet.cell(11, 8 + ks_inc, "Value")
            sheet.merge_cells(start_row=10, end_row=10, start_column=7 + ks_inc, end_column=8 + ks_inc)
            sheet.cell(10, 9 + ks_inc, "Purchase")
            self.ks_apply_style(sheet.cell(10, 9 + ks_inc), True, True, False, True)
            sheet.cell(11, 9 + ks_inc, "Qty")
            sheet.cell(11, 10 + ks_inc, "Value")
            sheet.merge_cells(start_row=10, end_row=10, start_column=9 + ks_inc, end_column=10 + ks_inc)
            sheet.cell(10, 11 + ks_inc, "Internal")
            self.ks_apply_style(sheet.cell(10, 11 + ks_inc), True, True, False, True)
            sheet.cell(11, 11 + ks_inc, "Qty")
            sheet.cell(11, 12 + ks_inc, "Value")
            sheet.merge_cells(start_row=10, end_row=10, start_column=11 + ks_inc, end_column=12 + ks_inc)
            sheet.cell(10, 13 + ks_inc, "Adjustment")
            self.ks_apply_style(sheet.cell(10, 13 + ks_inc), True, True, False, True)
            sheet.cell(11, 13 + ks_inc, "Qty")
            sheet.cell(11, 14 + ks_inc, "Value")
            sheet.merge_cells(start_row=10, end_row=10, start_column=13 + ks_inc, end_column=14 + ks_inc)
            sheet.cell(10, 15 + ks_inc, "Scrap")
            self.ks_apply_style(sheet.cell(10, 15 + ks_inc), True, True, False, True)
            sheet.cell(11, 15 + ks_inc, "Qty")
            sheet.cell(11, 16 + ks_inc, "Value")
            sheet.merge_cells(start_row=10, end_row=10, start_column=15 + ks_inc, end_column=16 + ks_inc)
            ks_inc += 10
        # else:
        #     if 'ks_sheet' in locals():
        #         sheet = ks_sheet[1]

        # get qty available
        self.env.cr.execute("""
        SELECT ks_product_id, ks_product_type,ks_product_categ_id,ks_product_name,ks_location_id,ks_company_id,
               ks_product_sales_price, ks_product_qty_available, ks_product_code
        FROM ks_warehouse_report 
            WHERE ks_company_id = %s and ks_product_qty_available != 0 and ks_usage = 'internal'
        order by ks_location_id
        """ % self.ks_company_id.id)

        datas = self.env.cr.fetchall()
        if not datas:
            raise ValidationError(_("Opps! There are no data."))

        # dates_in = self.ks_data_in_date()
        #
        # datas = self.ks_merge_data(datas, dates_in)
        #
        # datas = self.ks_apply_filter(datas)

        if datas:
            j = 1;
            i = 1;
            row = 12;
            col = 0
            for data in datas:
                self.ks_set_default_5_rows_left(sheet, row, j, data)
                ks_col = 0
                ks_cost = self.env['product.product'].browse(data[0]).product_tmpl_id.standard_price
                start = datetime.strptime(str(self.ks_date_from), "%Y-%m-%d")
                for i in range(0, (self.ks_date_to - self.ks_date_from).days, self.ks_duration):
                    # if 14 + ks_col > 255:  # It will enter the loop if column gretaer than 255 to split the sheet into new
                    #     ks_col = 0
                    #     if 'nd_sheet' in locals(): nd_sheet += 1
                    #     if 'nd_sheet' not in locals(): nd_sheet = 2
                    #     sheet = ks_sheet[nd_sheet]
                    #     self.ks_set_default_5_rows_left(sheet, row, j, data)

                    stop = start + relativedelta(days=period_length)
                    stop = ks_date_to if stop > ks_date_to else stop

                    pro_data = self.ks_get_pro_data(data, start, stop)

                    sheet.cell(row, 7 + ks_col, pro_data[0][0] if pro_data else 0)
                    sheet.cell(row, 8 + ks_col, (pro_data[0][0] if pro_data else 0) * ks_cost)
                    sheet.cell(row, 9 + ks_col, pro_data[0][1] if pro_data else 0)
                    sheet.cell(row, 10 + ks_col, (pro_data[0][1] if pro_data else 0) * ks_cost)
                    sheet.cell(row, 11 + ks_col, pro_data[0][2] if pro_data else 0)
                    sheet.cell(row, 12 + ks_col, (pro_data[0][2] if pro_data else 0) * ks_cost)
                    sheet.cell(row, 13 + ks_col, pro_data[0][3] if pro_data else 0)
                    sheet.cell(row, 14 + ks_col, (pro_data[0][3] if pro_data else 0) * ks_cost)
                    sheet.cell(row, 15 + ks_col, pro_data[0][4] if pro_data else 0)
                    sheet.cell(row, 16 + ks_col, (pro_data[0][4] if pro_data else 0) * ks_cost)
                    start = stop
                    ks_col += 10
                # else:
                #     if 'ks_sheet' in locals(): sheet = ks_sheet[1]
                #     if 'nd_sheet' in locals(): nd_sheet = 1

                row += 1
                i += 1
                j += 1
        output = StringIO()
        filename = (str(pathlib.Path(__file__).parent) + '/' + str(report_name) + '.xlsx')
        _log.info("Filepath: %s" % (filename))
        workbook.save(filename)
        fp = open(filename, "rb")
        file_data = fp.read()
        out = base64.encodebytes(file_data)

        # Files actions
        attach_vals = {
            'report_name': str(report_name) + '.xlsx',
            'datas': out,
        }

        act_id = self.env['ks.warehouse.report.ageing.with.movement.out'].create(attach_vals)
        fp.close()
        os.remove(filename)
        _log.info("File closed and removed.")
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'ks.warehouse.report.ageing.with.movement.out',
            'res_id': act_id.id,
            'view_type': 'form',
            'view_mode': 'form',
            'context': self.env.context,
            'target': 'new',
        }

    
    def ks_set_default_5_rows_left(self, sheet, row, j, data):
        sheet.cell(row, 1, j)
        sheet.cell(row, 2, data[0])
        if data[1] == 'product':
            sheet.cell(row, 3, 'Stockable')
        elif data[1] == 'consu':
            sheet.cell(row, 3, 'Consumable')
        if data[2]:
            catge_id = self.env['product.category'].browse(int(data[2]))
        sheet.cell(row, 4, catge_id.name)
        sheet.cell(row, 5, data[3])
        if data[4]:
            location_id = self.env['stock.location'].browse(int(data[4]))
        sheet.cell(row, 6, location_id.display_name)

    
    def ks_get_pro_data(self, data, start, stop):
        company = self.ks_company_id.id
        start = fields.Datetime.to_datetime(start)
        stop = fields.Datetime.to_datetime(stop)
        self.env.cr.execute("""
        with adjust as(
                select sml.product_id, sml.location_id, sm.company_id, 
                    sum(sml.qty_done) as adjustment,
                    sum(case when spt.code= 'internal' then sml.qty_done else 0 end) as internal_adjust
                from stock_move_line as sml
                    left join stock_move as sm on sm.id = sml.move_id
                    left join stock_location as sld on sld.id = sm.location_dest_id
                    left join stock_picking_type as spt on spt.id = sm.picking_type_id
                where sml.state = 'done' and sm.company_id = '%s' and sml.date between '%s' and '%s' and 
                    sml.product_id = '%s' and sml.location_id = '%s' and sld.scrap_location = False
                group by sml.product_id, sml.location_id, sm.company_id
             ),
             scrap as(
                    select scrap.product_id, scrap.location_id, sm.company_id, sum(scrap.scrap_qty) as scrapped
                    from stock_scrap as scrap
                        left join stock_move as sm on sm.id = scrap.move_id
                    where scrap.state = 'done' and sm.company_id = '%s' and scrap.date_done between '%s' and '%s'
                        and scrap.product_id = '%s' and scrap.location_id = '%s'
                    group by scrap.product_id, scrap.location_id, sm.company_id
                
             ),
             ad_sc as (
                 select adjust.product_id as p_id, adjust.location_id as p_lo, adjust.internal_adjust as p_ia, 
                    adjust.adjustment as p_ad, scrap.scrapped as p_sc 
                from adjust full outer join scrap 
                    on (adjust.product_id = scrap.product_id and adjust.location_id = scrap.location_id)
             ),   
             sale as (
                select sm.product_id, sm.location_id, sm.company_id,
                    sum(sm.product_uom_qty) as Sale_sum
                from stock_move as sm
                    left join stock_location as sl on sl.id = sm.location_id
                    left join stock_location as sld on sld.id = sm.location_dest_id
                where sm.state = 'done' and sm.company_id = '%s' and sm.date between '%s' and '%s'
                    and sl.usage in ('internal', 'transit') and sld.usage not in ('internal', 'transit')
                    and sm.product_id = '%s' and sm.location_id = '%s'
                    and sld.scrap_location = False
                group by sm.product_id, sm.location_id, sm.company_id
             ),
             purchase as (
                select sm.product_id, sm.location_id, sm.company_id,
                    sum(sm.product_uom_qty) as Purchase_sum
                from stock_move as sm
                    left join stock_location as sl on sl.id = sm.location_id
                    left join stock_location as sld on sld.id = sm.location_dest_id
                where sm.state = 'done' and sm.company_id = '%s' and sm.date between '%s' and '%s'
                    and sl.usage = 'supplier' and sld.usage in ('internal', 'transit')
                    and sm.product_id = '%s' and sm.location_id = '%s'
                group by sm.product_id, sm.location_id, sm.company_id
             ),
             sa_pu as (
                 select sale.product_id as s_id, sale.location_id as s_lo, sale.Sale_sum as s_sa, 
                    purchase.Purchase_sum as s_pu 
                from sale full outer join purchase 
                    on (sale.product_id = purchase.product_id and sale.location_id = purchase.location_id)
             )
             select case when sa_pu.s_sa is null then 0 else sa_pu.s_sa end, 
                    case when sa_pu.s_pu is null then 0 else sa_pu.s_pu end, 
                    case when ad_sc.p_ia is null then 0 else ad_sc.p_ia end, 
                    case when ad_sc.p_ad is null then 0 else ad_sc.p_ad end, 
                    case when ad_sc.p_sc is null then 0 else ad_sc.p_sc end 
             from ad_sc full outer join sa_pu on (sa_pu.s_id = ad_sc.p_id and sa_pu.s_lo = ad_sc.p_lo)
              """ % (
                4 * (company, start, stop, data[0], data[4])
            )
        )
        return self.env.cr.fetchall()


    class KSWarehouseReportAgeingWithMovementOUT(models.Model):
        _name = "ks.warehouse.report.ageing.with.movement.out"
        _description = "Stock Ageing No Movement report Out"

        datas = fields.Binary('File', readonly=True)
        report_name = fields.Char('Report Name', readonly=True)
