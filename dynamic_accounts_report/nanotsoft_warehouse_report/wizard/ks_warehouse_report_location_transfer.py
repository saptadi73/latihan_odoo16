from odoo import api, fields, models, _
from odoo.exceptions import ValidationError

import base64
from io import StringIO
import logging
import pathlib
import os

_log = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Border, Font, Alignment
except ImportError:
    _log.debug('Can not `import openpyxl`.')

class KSWarehouseReportlocationTransfers(models.Model):
    _name = "ks.warehouse.report.location.transfer"
    _description = "Stock location.transfers / Stock Report"

    ks_report = {'product_code': 0, 'product_type': 1, 'product_categ_id': 2, 'product_name': 3, 'location_id': 4,
                 'company_id': 5, 'product_sales_price': 6, 'product_qty_available': 7, 'product_id': 8}
    kr_adjustment= {'product_id': 0, 'location_id': 1, 'company_id': 2, 'date': 3, 'product_name': 4, 'product_type': 5,
                    'product_categ_id': 6}

    ks_name = fields.Char(default='Stock Location Transfer Report')
    ks_date_from = fields.Date('Start Date', required=True)
    ks_date_to = fields.Date('End Date', required=True)
    ks_company_id = fields.Many2one('res.company', 'Company', required=True,
                                 default=lambda self: self.env.user.company_id)

    def ks_action_generate_report(self):
        return True

    def ks_apply_style(self, ks_cell, kc='', vc='', sz=False, wp=False):
        ks_cell.alignment = Alignment(horizontal="center" if kc else '', vertical="center" if vc else '',
                                      wrap_text=wp)
        if sz: ks_cell.font = Font(b=True, size=sz)

    def ks_create_workbook_header(self, report_name, sheet):
        sheet.title = str(report_name)

        sheet['A1'] = str(report_name)
        self.ks_apply_style(sheet['A1'], True, True, 20, True)
        sheet.merge_cells(start_row=1, end_row=2, start_column=1, end_column=20)

        sheet['A3'] = "COMPANY : " + self.ks_company_id.name
        self.ks_apply_style(sheet['A3'], True, True, 14, True)
        sheet.merge_cells(start_row=3, end_row=3, start_column=1, end_column=20)

        sheet['A4'] = 'FROM : ' + str(self.ks_date_from) + ' | TO : ' + str(self.ks_date_to)
        self.ks_apply_style(sheet['A4'], True, True, 10, True)
        sheet.merge_cells(start_row=4, end_row=4, start_column=1, end_column=20)

        sheet['A6'] = "REPORT"
        self.ks_apply_style(sheet['A6'], True, True, 14, True)
        sheet.merge_cells(start_row=6, end_row=7, start_column=1, end_column=20)

        sheet['A8'] = "S.NO"
        sheet.merge_cells(start_row=8, end_row=9, start_column=1, end_column=1)

        sheet['B8'] = "Reference/Code"
        self.ks_apply_style(sheet['B8'], True, True, False, True)
        sheet.merge_cells(start_row=8, end_row=9, start_column=2, end_column=2)

        sheet['C8'] = "Type"
        sheet.merge_cells(start_row=8, end_row=9, start_column=3, end_column=3)

        sheet['D8'] = "Category"
        sheet.merge_cells(start_row=8, end_row=9, start_column=4, end_column=4)

        sheet['E8'] = "Product"
        sheet.merge_cells(start_row=8, end_row=9, start_column=5, end_column=5)

        sheet['F8'] = "Company"
        sheet.merge_cells(start_row=8, end_row=9, start_column=6, end_column=6)

        sheet['G8'] = "Transfers Dates"
        sheet['G9'] = "Last Sale"
        sheet['H9'] = "Last Purch"
        sheet['I9'] = "Last Adjusts"
        sheet.merge_cells(start_row=8, end_row=8, start_column=7, end_column=9)

        sheet['J8'] = "Transfers Location"
        sheet['J9'] = "Last Sale"
        sheet['K9'] = "Last Purch"
        sheet['L9'] = "Last Adjusts"
        sheet.merge_cells(start_row=8, end_row=8, start_column=10, end_column=12)

        sheet.freeze_panes = 'C10'
    
    def ks_generate_xlsx_report(self):
        report_name = self.ks_name
        workbook = openpyxl.Workbook()

        sheet = workbook.active

        self.ks_create_workbook_header(report_name, sheet)


        # get qty available
        # self.env.cr.execute("""
        # SELECT ks_product_code,ks_product_type,ks_product_categ_id,ks_product_name,ks_location_id,ks_company_id,
        #        ks_product_sales_price, ks_product_qty_available, ks_product_id
        # FROM ks_warehouse_report
        # WHERE ks_company_id = %s and ks_product_qty_available != 0
        # order by ks_location_id
        # """ % self.ks_company_id.id)
        #
        # datas = self.env.cr.fetchall()
        # if not datas:
        #     raise ValidationError(_("Opps! There are no data."))

        ks_adjusted_stock = self.ks_adjusted_stock()

        ks_sale_product = self.ks_sale_product()

        ks_purchase_product = self.ks_purchase_product()

        datas = self.ks_merge_data(ks_adjusted_stock, ks_sale_product, ks_purchase_product)

        # datas = self.ks_apply_filter(datas)

        if datas:
            i = 1;
            row = 10;
            col = 0
            for data in datas:
                sheet.cell(row, 1, i)
                sheet.cell(row, 2, data[0])
                if data[1] == 'product':
                    sheet.cell(row, 3, 'Stockable')
                elif data[1] == 'consu':
                    sheet.cell(row, 3, 'Consumable')
                if data[2]:
                    catge_id = self.env['product.category'].browse(int(data[2]))
                sheet.cell(row, 4, catge_id.name)
                sheet.cell(row, 5, data[3])
                # if data[4]:
                #     location_id = self.env['stock.location'].browse(int(data[4]))
                # sheet.write(row, 5, location_id.display_name, style0)
                # sheet.write(row, 5, 'WH', style0)
                if data[4]:
                    comp_id = self.env['res.company'].browse(int(data[4]))
                sheet.cell(row, 6, comp_id.name)
                sheet.cell(row, 7, data[5] if data[5] != 0 else '_')
                sheet.cell(row, 8, data[6] if data[6] != 0 else '_')
                sheet.cell(row, 9, data[7] if data[7] != 0 else '_')
                if data[8]:
                    s_location_id = self.env['stock.location'].browse(int(data[8]))
                    sheet.cell(row, 10, s_location_id.display_name if s_location_id else '' )
                else:
                    sheet.cell(row, 10, '_' )
                if data[9]:
                    p_location_id = self.env['stock.location'].browse(int(data[9]))
                    sheet.cell(row, 11, p_location_id.display_name if p_location_id else '' )
                else:
                    sheet.cell(row, 11, '_' )
                if data[10]:
                    a_location_id = self.env['stock.location'].browse(int(data[10]))
                    sheet.cell(row, 12, a_location_id.display_name if a_location_id else '' )
                else:
                    sheet.cell(row, 12, '_' )

                row += 1
                i += 1
        output = StringIO()
        filename = (str(pathlib.Path(__file__).parent) + '/' + str(report_name) + '.xlsx')
        _log.info("Filepath: %s" % (filename))
        workbook.save(filename)
        fp = open(filename, "rb")
        file_data = fp.read()
        out = base64.encodebytes(file_data)

        # Files actions
        attach_vals = {
            'report_name': str(report_name) + '.xlsx',
            'datas': out,
        }

        act_id = self.env['ks.warehouse.report.location.transfer.out'].create(attach_vals)
        fp.close()
        os.remove(filename)
        _log.info("File closed and removed.")
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'ks.warehouse.report.location.transfer.out',
            'res_id': act_id.id,
            'view_type': 'form',
            'view_mode': 'form',
            'context': self.env.context,
            'target': 'new',
        }

    
    def ks_purchase_product(self):
        ks_date_from = fields.Datetime.to_datetime(self.ks_date_from)  # not sl.usage not in ('internal', 'transit')
        self.env.cr.execute("""
            select sm.product_id, sm.location_dest_id, sm.company_id,
                max(sm.date)
            from stock_move as sm
                left join stock_location as sl on sl.id = sm.location_id
                left join stock_location as sld on sld.id = sm.location_dest_id
            where sm.state = 'done' and sm.company_id = '%s' and sm.date between '%s' and '%s' 
                and sl.usage = 'supplier' and sld.usage in ('internal', 'transit')
            group by sm.product_id, sm.location_dest_id, sm.company_id
        """ % (self.ks_company_id.id, ks_date_from, fields.Datetime.to_datetime(self.ks_date_to)))

        purchase_date = self.env.cr.fetchall()
        if not purchase_date:
            return {}
        else:
            ks_dict = dict()
            for ks in purchase_date:  # product_id + company_id : date, location_id
                id = str(ks[0]) + str(ks[2])
                if not ks_dict.get(id, 0):
                    ks_dict[id] = ks[3], ks[1]
                elif ks_dict.get(id)[0] < ks[3]:
                    ks_dict[id] = ks[3], ks[1]
            purchase_date = ks_dict
        return purchase_date

    
    def ks_sale_product(self):
        ks_date_from = fields.Datetime.to_datetime(self.ks_date_from)
        self.env.cr.execute("""
            select sm.product_id, sm.location_id, sm.company_id,
                max(sm.date)
            from stock_move as sm
                left join stock_location as sl on sl.id = sm.location_id
                left join stock_location as sld on sld.id = sm.location_dest_id
            where sm.state = 'done' and sm.company_id = '%s' and sm.date between '%s' and '%s' 
                and sl.usage in ('internal', 'transit') and sld.usage not in ('internal', 'transit')
                and sld.scrap_location = False
            group by sm.product_id, sm.location_id, sm.company_id
        """ % (self.ks_company_id.id, ks_date_from, fields.Datetime.to_datetime(self.ks_date_to)))

        sale_date = self.env.cr.fetchall()
        if not sale_date:
            return {}
        else:
            ks_dict = dict()
            for ks in sale_date:  # product_id + company_id : date, location_id
                id = str(ks[0]) + str(ks[2])
                if not ks_dict.get(id,0):
                    ks_dict[id] = ks[3], ks[1]
                elif ks_dict.get(id)[0] < ks[3]:
                    ks_dict[id] = ks[3], ks[1]
            sale_date = ks_dict
        return sale_date

    
    def ks_adjusted_stock(self):
        # get the stock_quant data via date in query
        ks_date_from = fields.Datetime.to_datetime(self.ks_date_from)
        self.env.cr.execute("""
            select sml.product_id, sml.location_id, sm.company_id, max(sml.date), 
                max(pt.name) as ks_product_name, max(pt.type) as ks_product_type, max(pc.id) as ks_product_categ_id 
            from stock_move_line as sml
                left join stock_move as sm on sm.id = sml.move_id
                left join stock_location as sld on sld.id = sm.location_dest_id
                LEFT JOIN product_product as pp ON pp.id = sml.product_id
                LEFT JOIN product_template as pt ON pt.id = pp.product_tmpl_id
                LEFT JOIN product_category as pc ON pc.id = pt.categ_id
            where sml.state = 'done' and sm.company_id = '%s' and sml.date between '%s' and '%s' 
                and sld.scrap_location = False
            group by sml.product_id, sml.location_id, sm.company_id
        """ % (self.ks_company_id.id, ks_date_from, fields.Datetime.to_datetime(self.ks_date_to)))

        adjusted_date = self.env.cr.fetchall()
        if not adjusted_date:
            raise ValidationError(_("Opps! There are no data."))
        return adjusted_date

    
    def ks_merge_data(self, adjusted, sale, purchase):
        ks_list, ks_prod = [], dict()
        kid = self.kr_adjustment
        for date in adjusted:
            ks_sale = sale.get(str(date[kid['product_id']]) + str(date[kid['company_id']]), (0,0))
            if ks_sale[0]:
                ks_sale = ks_sale[0].strftime('%d-%m-%Y'), ks_sale[1]
                del sale[str(date[kid['product_id']]) + str(date[kid['company_id']])]
            ks_purchase = purchase.get(str(date[kid['product_id']]) + str(date[kid['company_id']]), (0,0))
            if ks_purchase[0]:
                ks_purchase = ks_purchase[0].strftime('%d-%m-%Y'), ks_purchase[1]
                del purchase[str(date[kid['product_id']]) + str(date[kid['company_id']])]
            if not ks_prod.get(date[kid['product_id']], 0):
                ks_list.append(
                    (date[kid['product_id']], date[kid['product_type']], date[kid['product_categ_id']],
                     date[kid['product_name']] , date[kid['company_id']],
                     ks_sale[0], ks_purchase[0], date[kid['date']].strftime('%d-%m-%Y'),
                     ks_sale[1], ks_purchase[1], date[kid['location_id']])
                )
                ks_prod[date[kid['product_id']]] = len(ks_list) - 1, date[kid['date']]
            elif ks_prod.get(date[kid['product_id']])[1] < date[kid['date']]:
                index = ks_prod.get(date[kid['product_id']])[0]
                temp = list(ks_list[index])
                temp[7], temp[10] = date[kid['date']].strftime('%d-%m-%Y'),  date[kid['location_id']]
                ks_list[index] = tuple(temp)
        if not ks_list:
            raise ValidationError(_("Opps! There are no data."))
        return ks_list

    
    # def ks_apply_filter(self, data):
    #         # ks_data = self.ks_exhausted_filter(data)
    #         ks_data = data
    #         if not ks_data:
    #             raise ValidationError(_("Opps! There are no data."))
    #         return ks_data

    class KSWarehouseReportlocationTransferOUT(models.Model):
        _name = "ks.warehouse.report.location.transfer.out"
        _description = "Stock location Transfer report Out"

        datas = fields.Binary('File', readonly=True)
        report_name = fields.Char('Report Name', readonly=True)
