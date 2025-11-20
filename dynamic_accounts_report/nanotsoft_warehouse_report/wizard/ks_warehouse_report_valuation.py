from odoo import api, fields, models, _
from odoo.exceptions import ValidationError

import base64
from io import StringIO
import logging
import pathlib
import os

_log = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Border, Font, Alignment
except ImportError:
    _log.debug('Can not `import openpyxl`.')



class KSWarehouseReportValuation(models.Model):
    _name = "ks.warehouse.report.valuation"
    _description = "Stock Valuation report"
    # _auto = False

    ks_report = {'product_code':0, 'product_type':1, 'product_categ_id':2, 'product_name':3, 'location_id':4,
                 'company_id':5, 'product_sales_price':6, 'product_qty_available':7, 'product_id':8, 'product_barcode':9}
    kr_in_dates = {'product_id':0, 'location_id':1, 'company_id':2, 'opening_stock':3, 'closing_stock':4, 'qty_date':5}

    ks_name = fields.Char(default='Stock Valuation Report')
    ks_date_from = fields.Date('Start Date', required=True)
    ks_date_to = fields.Date('End Date', required=True)
    ks_inventory_loss = fields.Boolean('Include Inventory Loss')
    ks_company_id = fields.Many2one('res.company', 'Company', required=True,
                                 default=lambda self: self.env.user.company_id)
    ks_show_exhausted = fields.Boolean('Show Exhausted')
    ks_show_opening = fields.Boolean('Show Opening', default=True)
    ks_show_closing = fields.Boolean('Show Closing', default=True)
    ks_show_adjustment = fields.Boolean('Show Adjustment', default=True)
    ks_show_scrap_loss = fields.Boolean('Show Scrap/Loss', default=True)
    ks_show_current = fields.Boolean('Show Current', default=True)

    def ks_apply_style(self, ks_cell, kc='', vc='', sz=False, wp=False):
        ks_cell.alignment = Alignment(horizontal="center" if kc else '', vertical="center" if vc else '',
                                      wrap_text=wp)
        if sz: ks_cell.font = Font(b=True, size=sz)

    def ks_create_workbook_header(self, report_name, sheet):
        sheet.title = str(report_name)

        sheet['A1'] = str(report_name)
        self.ks_apply_style(sheet['A1'], True, True, 20, True)
        sheet.merge_cells(start_row=1, end_row=2, start_column=1, end_column=20)

        sheet['A3'] = "COMPANY : " + self.ks_company_id.name
        self.ks_apply_style(sheet['A3'], True, True, 14, True)
        sheet.merge_cells(start_row=3, end_row=3, start_column=1, end_column=20)

        sheet['A4'] = 'FROM : ' + str(self.ks_date_from) + ' | TO : ' + str(self.ks_date_to)
        self.ks_apply_style(sheet['A4'], True, True, 10, True)
        sheet.merge_cells(start_row=4, end_row=4, start_column=1, end_column=20)

        sheet['A6'] = "REPORT"
        self.ks_apply_style(sheet['A6'], True, True, 14, True)
        sheet.merge_cells(start_row=6, end_row=7, start_column=1, end_column=20)

        sheet['A8'] = "S.NO"
        sheet.merge_cells(start_row=8, end_row=9, start_column=1, end_column=1)

        sheet['B8'] = "Reference/Code"
        self.ks_apply_style(sheet['B8'], True, True, False, True)
        sheet.merge_cells(start_row=8, end_row=9, start_column=2, end_column=2)

        sheet['C8'] = "Barcode"
        sheet.merge_cells(start_row=8, end_row=9, start_column=3, end_column=3)

        sheet['D8'] = "Type"
        sheet.merge_cells(start_row=8, end_row=9, start_column=4, end_column=4)

        sheet['E8'] = "Category"
        sheet.merge_cells(start_row=8, end_row=9, start_column=5, end_column=5)

        sheet['F8'] = "Product"
        sheet.merge_cells(start_row=8, end_row=9, start_column=6, end_column=6)

        sheet['G8'] = "Location"
        sheet.merge_cells(start_row=8, end_row=9, start_column=7, end_column=7)

        sheet['H8'] = "Company"
        sheet.merge_cells(start_row=8, end_row=9, start_column=8, end_column=8)

        sheet['I8'] = "Available Qty"
        self.ks_apply_style(sheet['H8'], True, True, False, True)
        sheet.merge_cells(start_row=8, end_row=9, start_column=9, end_column=9)

        sheet['J8'] = "Cost"
        sheet.merge_cells(start_row=8, end_row=9, start_column=10, end_column=10)

        sheet['K8'] = "Sales Price"
        sheet.merge_cells(start_row=8, end_row=9, start_column=11, end_column=11)

        self.ks_dynamic_sheet(sheet)

        sheet.freeze_panes = 'C10'

    def ks_generate_xlsx_report(self):
        report_name = self.ks_name
        workbook = openpyxl.Workbook()

        sheet = workbook.active

        self.ks_create_workbook_header(report_name, sheet)

        # get qty available
        if self.ks_inventory_loss:
            self.env.cr.execute("""
            SELECT ks_product_code,ks_product_type,ks_product_categ_id,ks_product_name,ks_location_id,ks_company_id,
                   ks_product_sales_price, ks_product_qty_available,ks_product_id,ks_product_barcode
            FROM ks_warehouse_report 
            WHERE ks_company_id = %s and ks_product_qty_available != 0 and (ks_usage = 'internal' or ks_usage = 'inventory')
                order by ks_location_id
            """ % self.ks_company_id.id)
        else:
            self.env.cr.execute("""
                    SELECT ks_product_code,ks_product_type,ks_product_categ_id,ks_product_name,ks_location_id,ks_company_id,
                           ks_product_sales_price, ks_product_qty_available,ks_product_id,ks_product_barcode
                    FROM ks_warehouse_report 
                    WHERE ks_company_id = %s and ks_product_qty_available != 0 and ks_usage = 'internal' 
                        order by ks_location_id
                    """ % self.ks_company_id.id)

        datas = self.env.cr.fetchall()
        if not datas:
            raise ValidationError(_("Opps! There are no data."))

        dates_in = self.ks_data_in_date()

        ks_adjusted_stock = self.ks_adjusted_stock()

        ks_scrap_stock = self.ks_scrap_stock()

        datas = self.ks_merge_data(datas, dates_in, ks_adjusted_stock, ks_scrap_stock)


        if datas:
            i = 1; row = 10; col = 0
            for data in datas:
                # if (not self.location_id or self.location.id == data[4]) or (not self.location_id or self.location.id == data[4])
                sheet.cell(row, 1, i)
                sheet.cell(row, 2, data[0])
                sheet.cell(row, 3, data[19])
                if data[1] == 'product':
                    sheet.cell(row, 4, 'Stockable')
                elif data[1]  == 'consu':
                    sheet.cell(row, 4, 'Consumable')
                if data[2]:
                    catge_id = self.env['product.category'].browse(int(data[2]))
                sheet.cell(row, 5, catge_id.name)
                sheet.cell(row, 6, data[3])
                if data[4]:
                    location_id = self.env['stock.location'].browse(int(data[4]))
                sheet.cell(row, 7, location_id.display_name)
                if data[5]:
                    comp_id = self.env['res.company'].browse(int(data[5]))
                sheet.cell(row, 8, comp_id.name)
                sheet.cell(row, 9, data[6])
                sheet.cell(row, 10, data[7])
                sheet.cell(row, 11, data[8])
                c_1, c_2 = 12, 13
                if self.ks_show_opening:
                    sheet.cell(row, c_1, data[9])
                    sheet.cell(row, c_2, data[10])
                    c_1, c_2 = c_1 + 2, c_2 + 2
                if self.ks_show_closing:
                    sheet.cell(row, c_1, data[11])
                    sheet.cell(row, c_2, data[12])
                    c_1, c_2 = c_1 + 2, c_2 + 2
                if self.ks_show_adjustment:
                    sheet.cell(row, c_1, data[13])
                    sheet.cell(row, c_2, data[14])
                    c_1, c_2 = c_1 + 2, c_2 + 2
                if self.ks_show_scrap_loss:
                    sheet.cell(row, c_1, data[15])
                    sheet.cell(row, c_2, data[16])
                    c_1, c_2 = c_1 + 2, c_2 + 2
                if self.ks_show_current:
                    sheet.cell(row, c_1, data[17])
                    sheet.cell(row, c_2, data[18])

                row += 1
                i += 1
        output = StringIO()
        filename = (str(pathlib.Path(__file__).parent) + '/' + str(report_name) + '.xlsx')
        _log.info("Filepath: %s" % (filename))
        workbook.save(filename)
        fp = open(filename, "rb")
        file_data = fp.read()
        out = base64.encodebytes(file_data)

        # Files actions
        attach_vals = {
            'report_name': str(report_name) + '.xlsx',
            'datas': out,
        }

        act_id = self.env['ks.warehouse.report.valuation.out'].create(attach_vals)
        fp.close()
        os.remove(filename)
        _log.info("File closed and removed.")
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'ks.warehouse.report.valuation.out',
            'res_id': act_id.id,
            'view_type': 'form',
            'view_mode': 'form',
            'context': self.env.context,
            'target': 'new',
        }


    def ks_action_generate_report(self):
        return True


    def ks_dynamic_sheet(self, sheet):
        c_1, c_2 = 12, 13
        if self.ks_show_opening:
            sheet.cell(9, c_1, "Qty.")
            sheet.cell(9, c_2, "Value")
            sheet.cell(8, c_1, "Opening Stock")
            self.ks_apply_style(sheet.cell(8, c_1), True, True, False, True)
            sheet.merge_cells(start_row=8, end_row=8, start_column=c_1, end_column=c_2)
            c_1, c_2 = c_1 + 2, c_2 + 2
        if self.ks_show_closing:
            sheet.cell(9, c_1, "Qty.")
            sheet.cell(9, c_2, "Value")
            sheet.cell(8, c_1, "Closing Stock")
            self.ks_apply_style(sheet.cell(8, c_1), True, True, False, True)
            sheet.merge_cells(start_row=8, end_row=8, start_column=c_1, end_column=c_2)
            c_1, c_2 = c_1 + 2, c_2 + 2
        if self.ks_show_adjustment:
            sheet.cell(9, c_1, "Qty.")
            sheet.cell(9, c_2, "Value")
            sheet.cell(8, c_1, "Adjustment Stock")
            self.ks_apply_style(sheet.cell(8, c_1), True, True, False, True)
            sheet.merge_cells(start_row=8, end_row=8, start_column=c_1, end_column=c_2)
            c_1, c_2 = c_1 + 2, c_2 + 2
        if self.ks_show_scrap_loss:
            sheet.cell(9, c_1, "Qty.")
            sheet.cell(9, c_2, "Value")
            sheet.cell(8, c_1, "Scrap/Loss Stock")
            self.ks_apply_style(sheet.cell(8, c_1), True, True, False, True)
            sheet.merge_cells(start_row=8, end_row=8, start_column=c_1, end_column=c_2)
            c_1, c_2 = c_1 + 2, c_2 + 2
        if self.ks_show_current:
            sheet.cell(9, c_1, "Qty.")
            sheet.cell(9, c_2, "Value")
            sheet.cell(8, c_1, "Stock In Hand")
            self.ks_apply_style(sheet.cell(8, c_1), True, True, False, True)
            sheet.merge_cells(start_row=8, end_row=8, start_column=c_1, end_column=c_2)


    def ks_scrap_stock(self):
        # get the stock_quant data via date in query
        ks_date_from = fields.Datetime.to_datetime(self.ks_date_from)
        self.env.cr.execute("""
            select scrap.product_id, scrap.location_id, sm.company_id, sum(scrap.scrap_qty) 
            from stock_scrap as scrap
                left join stock_move as sm on sm.id = scrap.move_id
            where scrap.state = 'done' and sm.company_id = '%s' and scrap.date_done between '%s' and '%s'
            group by scrap.product_id, scrap.location_id, sm.company_id
        """ % (self.ks_company_id.id, ks_date_from, fields.Datetime.to_datetime(self.ks_date_to)))

        scrap_date = self.env.cr.fetchall()
        if not scrap_date:
            return {}
        else:
            ks_dict = dict()
            for ks in scrap_date: # product_id + location_id + company_id : qty_done(state=done)
                ks_dict[str(ks[0])+str(ks[1])+str(ks[2])] = ks[3]
            scrap_date = ks_dict
        return scrap_date


    def ks_adjusted_stock(self):
        # get the stock_quant data via date in query
        ks_date_from = fields.Datetime.to_datetime(self.ks_date_from)
        self.env.cr.execute("""
            select sml.product_id, sml.location_id, sm.company_id, sum(sml.qty_done) 
            from stock_move_line as sml
                left join stock_move as sm on sm.id = sml.move_id
                left join stock_location as sld on sld.id = sm.location_dest_id
            where sml.state = 'done' and sm.company_id = '%s' and sml.date between '%s' and '%s' 
                and sld.scrap_location = False
            group by sml.product_id, sml.location_id, sm.company_id
        """ % (self.ks_company_id.id, ks_date_from, fields.Datetime.to_datetime(self.ks_date_to)))

        adjusted_date = self.env.cr.fetchall()
        if not adjusted_date:
            return {}
        else:
            ks_dict = dict()
            for ks in adjusted_date: # product_id + location_id + company_id : qty_done(state=done)
                ks_dict[str(ks[0])+str(ks[1])+str(ks[2])] = ks[3]
            adjusted_date = ks_dict
        return adjusted_date


    def ks_data_in_date(self):
        # get the stock_quant data via date in query
        ks_date_from = fields.Datetime.to_datetime(self.ks_date_from)
        _select = """ select sq.product_id, sm.location_dest_id, sq.company_id,
            sum(case when sq.create_date < '%s' then sq.quantity else 0 end) as opening_stock,
            sum(sq.quantity) as closing_stock,
            sum(case when sq.create_date >= '%s' then sq.quantity else 0 end) as qty_date 
        """
        _from = """ from stock_valuation_layer as sq
            left join stock_move as sm on sm.id = sq.stock_move_id
            left join stock_location as sld on sld.id = sm.location_dest_id
	        left join stock_location as sl ON sl.id = sm.location_id 
        """
        _where = """ where sq.company_id = '%s'
            and sq.create_date <= '%s' 
        """
        _groupby = """ group by sq.product_id, sm.location_dest_id, sq.company_id 
        """
        self.env.cr.execute((_select + _from + _where + _groupby)
                % (ks_date_from, ks_date_from, self.ks_company_id.id, fields.Datetime.to_datetime(self.ks_date_to)))

        dates_in = self.env.cr.fetchall()
        if not dates_in:
            raise ValidationError(_("Opps! There are no data."))
        return dates_in


    def ks_merge_data(self, datas, dates_in, adjusted={}, scrap={}):
        ks_list = []
        kr = self.ks_report
        kid = self.kr_in_dates
        for date in dates_in:
            for data in datas:
                dp_id, dl_id, dc_id= data[kr['product_id']], data[kr['location_id']], data[kr['company_id']]
                if dp_id == date[kid['product_id']] and dl_id == date[kid['location_id']] and dc_id == date[kid['company_id']]:
                    ks_cost = self.env['product.product'].browse(date[kid['product_id']]).product_tmpl_id.standard_price
                    ks_adjusted = adjusted.get(str(dp_id)+str(dl_id)+str(dc_id), 0)
                    ks_scrap = scrap.get(str(dp_id)+str(dl_id)+str(dc_id), 0)
                    ks_qty_available = date[kid['qty_date']]
                    if not self.ks_show_exhausted:
                        if ks_qty_available < 0: ks_qty_available = 'not_allowed'
                    if ks_qty_available != 'not_allowed':
                        ks_data = 0
                        for rec in ks_list:
                            if rec[0] == data[kr['product_id']]:
                                ks_data = 1
                        if not ks_data:
                            ks_list.append(
                                (data[kr['product_id']], data[kr['product_type']], data[kr['product_categ_id']],
                                 data[kr['product_name']], data[kr['location_id']], data[kr['company_id']],
                                 date[kid['qty_date']], ks_cost, data[kr['product_sales_price']],
                                 date[kid['opening_stock']],
                                 date[kid['opening_stock']] * ks_cost, date[kid['closing_stock']],
                                 date[kid['closing_stock']] * ks_cost, ks_adjusted, ks_adjusted * ks_cost,
                                 ks_scrap, ks_scrap * ks_cost,
                                 date[kid['closing_stock']] - ks_adjusted-ks_scrap,
                                 (date[kid['closing_stock']] - ks_adjusted-ks_scrap) * ks_cost, data[kr['product_barcode']]

                                 )
                            )
        if not ks_list:
            raise ValidationError(_("Opps! There are no data."))
        return ks_list


    # def ks_apply_filter(self, data):
    #     ks_data = self.ks_exhausted_filter(data)
    #     if not ks_data:
    #         raise ValidationError(_("Opps! There are no data."))
    #     return ks_data


    # def ks_exhausted_filter(self, data):
    #     ks_data = []
    #     if not self.ks_show_exhausted:
    #         for ks in data:
    #             if ks[6] > 0:  # Ask if product type stockable filter is to be used?
    #                 ks_data.append(ks)
    #     else:
    #         ks_data = data
    #     return ks_data

class KSWarehouseReportValuationOUT(models.Model):
    _name = "ks.warehouse.report.valuation.out"
    _description = "Stock Valuation report Out"

    datas = fields.Binary('File', readonly=True)
    report_name = fields.Char('Report Name', readonly=True)
