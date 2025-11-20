from odoo import api, fields, models, _
from odoo.exceptions import ValidationError

import base64
from io import StringIO
import logging
import pathlib
import os

_log = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Border, Font, Alignment
except ImportError:
    _log.debug('Can not `import openpyxl`.')

class KSWarehouseReportstatuss(models.Model):
    _name = "ks.warehouse.report.status"
    _description = "Stock statuss / Stock Report"

    ks_report = {'product_code': 0, 'product_type': 1, 'product_categ_id': 2, 'product_name': 3, 'location_id': 4,
                 'company_id': 5, 'product_sales_price': 6, 'product_qty_available': 7, 'product_id': 8}
    ks_transfer = {'product_id': 0, 'location_id': 1, 'company_id': 2, 'in': 3, 'out': 4,
                  'in_refund': 5, 'out_refund': 6, 'status': 7, 'product_name': 8, 'product_type': 9,
                   'product_categ_id': 10}

    ks_name = fields.Char(default='Stock Status Report')
    ks_date_from = fields.Date('Start Date', required=True)
    ks_date_to = fields.Date('End Date', required=True)
    ks_company_id = fields.Many2one('res.company', 'Company', required=True,
                                 default=lambda self: self.env.user.company_id)
    
    def ks_action_generate_report(self):
        return True

    def ks_apply_style(self, ks_cell, kc='', vc='', sz=False, wp=False):
        ks_cell.alignment = Alignment(horizontal="center" if kc else '', vertical="center" if vc else '',
                                      wrap_text=wp)
        if sz: ks_cell.font = Font(b=True, size=sz)

    def ks_create_workbook_header(self, report_name, sheet):
        sheet.title = str(report_name)

        sheet['A1'] = str(report_name)
        self.ks_apply_style(sheet['A1'], True, True, 20, True)
        sheet.merge_cells(start_row=1, end_row=2, start_column=1, end_column=20)

        sheet['A3'] = "COMPANY : " + self.ks_company_id.name
        self.ks_apply_style(sheet['A3'], True, True, 14, True)
        sheet.merge_cells(start_row=3, end_row=3, start_column=1, end_column=20)

        sheet['A4'] = 'FROM : ' + str(self.ks_date_from) + ' | TO : ' + str(self.ks_date_to)
        self.ks_apply_style(sheet['A4'], True, True, 10, True)
        sheet.merge_cells(start_row=4, end_row=4, start_column=1, end_column=20)

        sheet['A6'] = "REPORT"
        self.ks_apply_style(sheet['A6'], True, True, 14, True)
        sheet.merge_cells(start_row=6, end_row=7, start_column=1, end_column=20)

        sheet['A8'] = "S.NO"
        sheet.merge_cells(start_row=8, end_row=9, start_column=1, end_column=1)

        sheet['B8'] = "Reference/Code"
        self.ks_apply_style(sheet['B8'], True, True, False, True)
        sheet.merge_cells(start_row=8, end_row=9, start_column=2, end_column=2)

        sheet['C8'] = "Type"
        sheet.merge_cells(start_row=8, end_row=9, start_column=3, end_column=3)

        sheet['D8'] = "Category"
        sheet.merge_cells(start_row=8, end_row=9, start_column=4, end_column=4)

        sheet['E8'] = "Product"
        sheet.merge_cells(start_row=8, end_row=9, start_column=5, end_column=5)

        sheet['F8'] = "Location"
        sheet.merge_cells(start_row=8, end_row=9, start_column=6, end_column=6)

        sheet['G8'] = "Company"
        sheet.merge_cells(start_row=8, end_row=9, start_column=7, end_column=7)

        sheet['H8'] = "Transfers"
        sheet['H9'] = "In Qty"
        sheet['I9'] = "Out Qty"
        sheet.merge_cells(start_row=8, end_row=8, start_column=8, end_column=9)

        sheet['J8'] = "Return Transfers"
        sheet['J9'] = "Return In"
        sheet['K9'] = "Return Out"
        sheet.merge_cells(start_row=8, end_row=8, start_column=10, end_column=11)

        sheet['L8'] = "Status"
        sheet.merge_cells(start_row=8, end_row=9, start_column=12, end_column=12)

        sheet.freeze_panes = 'C10'

    def ks_generate_xlsx_report(self):
        report_name = self.ks_name
        workbook = openpyxl.Workbook()

        sheet = workbook.active

        self.ks_create_workbook_header(report_name, sheet)

        # # get qty available
        # self.env.cr.execute("""
        # SELECT ks_product_code,ks_product_type,ks_product_categ_id,ks_product_name,ks_location_id,ks_company_id,
        #        ks_product_sales_price, ks_product_qty_available, ks_product_id
        # FROM ks_warehouse_report
        # WHERE ks_company_id = %s and ks_product_qty_available != 0
        # order by ks_location_id
        # """ % self.ks_company_id.id)
        #
        # datas = self.env.cr.fetchall()
        # if not datas:
        #     raise ValidationError(_("Opps! There are no data."))

        transfer = self.ks_transfers()

        datas = self.ks_merge_data(transfer)

        # datas = self.ks_apply_filter(datas)

        if datas:
            i = 1;
            row = 10;
            col = 0
            for data in datas:
                sheet.cell(row, 1, i)
                sheet.cell(row, 2, data[0])
                if data[1] == 'product':
                    sheet.cell(row, 3, 'Stockable')
                elif data[1] == 'consu':
                    sheet.cell(row, 3, 'Consumable')
                if data[2]:
                    catge_id = self.env['product.category'].browse(int(data[2]))
                sheet.cell(row, 4, catge_id.name)
                sheet.cell(row, 5, data[3])
                if data[4]:
                    location_id = self.env['stock.location'].browse(int(data[4]))
                sheet.cell(row, 6, location_id.display_name)
                if data[5]:
                    comp_id = self.env['res.company'].browse(int(data[5]))
                sheet.cell(row, 7, comp_id.name)
                sheet.cell(row, 8, data[6])
                sheet.cell(row, 9, data[7])
                sheet.cell(row, 10, data[8])
                sheet.cell(row, 11, data[9])
                sheet.cell(row, 12, data[10])

                row += 1
                i += 1
        output = StringIO()
        filename = (str(pathlib.Path(__file__).parent) + '/' + str(report_name) + '.xlsx')
        _log.info("Filepath: %s" % (filename))
        workbook.save(filename)
        fp = open(filename, "rb")
        file_data = fp.read()
        out = base64.encodebytes(file_data)

        # Files actions
        attach_vals = {
            'report_name': str(report_name) + '.xlsx',
            'datas': out,
        }

        act_id = self.env['ks.warehouse.report.status.out'].create(attach_vals)
        fp.close()
        os.remove(filename)
        _log.info("File closed and removed.")
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'ks.warehouse.report.status.out',
            'res_id': act_id.id,
            'view_type': 'form',
            'view_mode': 'form',
            'context': self.env.context,
            'target': 'new',
        }

    
    def ks_transfers(self):
        ks_date_from = fields.Datetime.to_datetime(self.ks_date_from)
        _select = """
            sm.product_id, sm.location_id, sm.company_id,
            	sum(case when spt.code = 'incoming' and sm.origin_returned_move_id is null and sm.inventory_id is null then sm.product_uom_qty else 0 end) as in,
                sum(case when spt.code = 'outgoing' and sm.origin_returned_move_id is null and sm.inventory_id is null then sm.product_uom_qty else 0 end) as out,
                sum(case when spt.code = 'incoming' and not sm.origin_returned_move_id is null and sm.inventory_id is null then sm.product_uom_qty else 0 end) as in_refund,
                sum(case when spt.code = 'outgoing' and not sm.origin_returned_move_id is null and sm.inventory_id is null then sm.product_uom_qty else 0 end) as out_refund,
                max(sm.state) as status, max(pt.name) as ks_product_name, max(pt.type) as ks_product_type, max(pc.id) as ks_product_categ_id
        """
        _from = """
            stock_move as sm
            left join stock_picking as sp on sp.id = sm.picking_id
            left join stock_picking_type as spt on spt.id = sm.picking_type_id
            LEFT JOIN product_product as pp ON pp.id = sm.product_id
            LEFT JOIN product_template as pt ON pt.id = pp.product_tmpl_id
            LEFT JOIN product_category as pc ON pc.id = pt.categ_id
        """
        _where = """
            sm.state = 'done' and sm.company_id = '%s' and 
            sp.scheduled_date between '%s' and '%s'
        """ % (self.ks_company_id.id, ks_date_from, fields.Datetime.to_datetime(self.ks_date_to))
        _groupby = """
            sm.product_id, sm.location_id, sm.company_id
        """
        self.env.cr.execute("""SELECT %s FROM %s WHERE %s GROUP BY %s
            """ % (_select, _from, _where, _groupby)
        )

        transfer = self.env.cr.fetchall()
        if not transfer:
            raise ValidationError(_("Opps! There are no data."))
        return transfer

    
    def ks_merge_data(self, transfer):
        ks_list = []
        kt = self.ks_transfer
        for date in transfer:
            # for data in datas:
            #     if data[kr['product_id']] == date[kt['product_id']] \
            #             and data[kr['company_id']] == date[kt['company_id']]:
                    ks_list.append(
                        (date[kt['product_id']], date[kt['product_type']], date[kt['product_categ_id']],
                         date[kt['product_name']], date[kt['location_id']], date[kt['company_id']],
                         date[kt['in']], date[kt['out']], date[kt['in_refund']], date[kt['out_refund']],
                         date[kt['status']]
                         )
                    )
        if not ks_list:
            raise ValidationError(_("Opps! There are no data."))
        return ks_list

    # def ks_apply_filter(self, data):
    #         # ks_data = self.ks_exhausted_filter(data)
    #         ks_data = data
    #         if not ks_data:
    #             raise ValidationError(_("Opps! There are no data."))
    #         return ks_data

    class KSWarehouseReportstatusOUT(models.Model):
        _name = "ks.warehouse.report.status.out"
        _description = "Stock status report Out"

        datas = fields.Binary('File', readonly=True)
        report_name = fields.Char('Report Name', readonly=True)
